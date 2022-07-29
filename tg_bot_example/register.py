import openpyxl


def create_user(telegram_id, name, email):
    book = openpyxl.load_workbook('user.xlsx')

    sheet = book.active

    sheet.append((telegram_id, name, email))

    book.save('user.xlsx')


def is_user_exists(telegram_id):
    book = openpyxl.load_workbook('user.xlsx')

    sheet = book.active

    row = 2
    while True:
        tg_id = sheet.cell(row=row, column=1).value
        if tg_id is None:
            return False
        if str(tg_id) == str(telegram_id):
            return True


if __name__ == "__main__":
    book = openpyxl.load_workbook('user.xlsx')

    sheet = book.active

    a1 = sheet['A1']

    print(a1.value)
